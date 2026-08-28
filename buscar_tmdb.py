import os
import requests
import openpyxl
from dotenv import load_dotenv

load_dotenv()

token = os.getenv("TMDB_TOKEN")

headers = {"Authorization": f"Bearer {token}", "accept": "application/json"}

archivo_excel = "datos/marvel_db.xlsx"

# Abrir Excel
libro = openpyxl.load_workbook(archivo_excel)
hoja = libro.active

# Buscar las columnas por nombre
encabezados = {}

for celda in hoja[1]:
    encabezados[celda.value] = celda.column

col_id = encabezados["id_tmdb"]
col_titulo = encabezados["titulo"]

# Recorrer películas
for fila in range(2, hoja.max_row + 1):

    titulo = hoja.cell(fila, col_titulo).value

    if not titulo:
        continue

    print(f"Buscando: {titulo}")

    url = "https://api.themoviedb.org/3/search/multi"

    params = {"query": titulo, "language": "es-ES"}

    respuesta = requests.get(url, headers=headers, params=params)

    if respuesta.status_code != 200:
        print("  ❌ Error:", respuesta.status_code)
        continue

    resultados = respuesta.json().get("results", [])

    # Buscar el primer resultado que sea película o serie
    resultado_valido = None

    for resultado in resultados:

        if resultado.get("media_type") in ["movie", "tv"]:
            resultado_valido = resultado
            break

    if resultado_valido:

        tmdb_id = resultado_valido["id"]

        hoja.cell(fila, col_id).value = tmdb_id

        print(
            f"  ✅ {resultado_valido.get('title') or resultado_valido.get('name')} "
            f"→ {tmdb_id}"
        )

    else:

        print("  ❌ No encontrado")


# Guardar Excel
libro.save(archivo_excel)

print("\nExcel actualizado correctamente.")
